"""
Utilities for PDF and Excel export
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFExporter:
    """
    Utility class for generating PDF reports
    """
    
    @staticmethod
    def create_header(canvas, doc):
        """Create PDF header with ZeroTec branding"""
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 16)
        canvas.drawString(inch, 10.5 * inch, "ZeroTec ERP")
        canvas.setFont('Helvetica', 10)
        canvas.drawString(inch, 10.3 * inch, "Zero Complicação, Total Gestão")
        canvas.line(inch, 10.2 * inch, 7.5 * inch, 10.2 * inch)
        canvas.restoreState()
    
    @staticmethod
    def create_footer(canvas, doc):
        """Create PDF footer with page number"""
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        page_num = canvas.getPageNumber()
        text = f"Página {page_num} - Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        canvas.drawRightString(7.5 * inch, 0.5 * inch, text)
        canvas.restoreState()
    
    @staticmethod
    def generate_dashboard_pdf(data):
        """
        Generate PDF for dashboard data
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Dashboard ZeroTec", title_style))
        elements.append(Spacer(1, 12))
        
        # Date
        date_style = styles['Normal']
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style))
        elements.append(Spacer(1, 20))
        
        # Vendas Section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#FF6B35'),
            spaceAfter=12
        )
        elements.append(Paragraph("VENDAS DO MES", section_style))
        
        vendas_data = [
            ['Métrica', 'Valor'],
            ['Total do Mês', f"R$ {data['vendas']['total_mes']:,.2f}"],
            ['Quantidade', str(data['vendas']['quantidade_mes'])],
            ['Ticket Médio', f"R$ {data['vendas']['ticket_medio']:,.2f}"],
            ['Crescimento', f"{data['vendas']['crescimento_percentual']:.1f}%"]
        ]
        
        vendas_table = Table(vendas_data, colWidths=[4*inch, 2.5*inch])
        vendas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(vendas_table)
        elements.append(Spacer(1, 20))
        
        # Financeiro Section
        elements.append(Paragraph("SITUACAO FINANCEIRA", section_style))
        
        financeiro_data = [
            ['Métrica', 'Valor'],
            ['Contas a Receber', f"R$ {data['financeiro']['saldo_contas_receber']:,.2f}"],
            ['Contas a Pagar', f"R$ {data['financeiro']['saldo_contas_pagar']:,.2f}"],
            ['Saldo Líquido', f"R$ {data['financeiro']['saldo_liquido']:,.2f}"],
            ['Vencidas (Receber)', f"R$ {data['financeiro']['contas_receber_vencidas']:,.2f}"],
            ['Vencidas (Pagar)', f"R$ {data['financeiro']['contas_pagar_vencidas']:,.2f}"]
        ]
        
        fin_table = Table(financeiro_data, colWidths=[4*inch, 2.5*inch])
        fin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(fin_table)
        elements.append(Spacer(1, 20))
        
        # CRM Section
        elements.append(Paragraph("CRM - PIPELINE DE VENDAS", section_style))
        
        crm_data = [
            ['Métrica', 'Valor'],
            ['Oportunidades Abertas', str(data['crm']['oportunidades_abertas'])],
            ['Valor do Pipeline', f"R$ {data['crm']['valor_pipeline']:,.2f}"]
        ]
        
        crm_table = Table(crm_data, colWidths=[4*inch, 2.5*inch])
        crm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(crm_table)
        
        
        # Build PDF with footer
        doc.build(elements, onFirstPage=PDFExporter.create_footer, onLaterPages=PDFExporter.create_footer)
        
        buffer.seek(0)
        return buffer


class ExcelExporter:
    """
    Utility class for generating Excel reports
    """
    
    @staticmethod
    def style_header(cell):
        """Apply header styling to cell"""
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def style_cell(cell):
        """Apply standard cell styling"""
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    @staticmethod
    def generate_dashboard_excel(data):
        """
        Generate Excel for dashboard data
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        
        # Title
        ws['A1'] = "ZeroTec ERP - Dashboard"
        ws['A1'].font = Font(bold=True, size=16, color="0066CC")
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Vendas Section
        row = 4
        ws[f'A{row}'] = "VENDAS DO MÊS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FF6B35")
        
        row += 1
        headers = ['Métrica', 'Valor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            ExcelExporter.style_header(cell)
        
        vendas_data = [
            ['Total do Mês', f"R$ {data['vendas']['total_mes']:,.2f}"],
            ['Quantidade', data['vendas']['quantidade_mes']],
            ['Ticket Médio', f"R$ {data['vendas']['ticket_medio']:,.2f}"],
            ['Crescimento', f"{data['vendas']['crescimento_percentual']:.1f}%"]
        ]
        
        for item in vendas_data:
            row += 1
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col, value=value)
                ExcelExporter.style_cell(cell)
        
        # Financeiro Section
        row += 2
        ws[f'A{row}'] = "SITUAÇÃO FINANCEIRA"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FF6B35")
        
        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            ExcelExporter.style_header(cell)
        
        fin_data = [
            ['Contas a Receber', f"R$ {data['financeiro']['saldo_contas_receber']:,.2f}"],
            ['Contas a Pagar', f"R$ {data['financeiro']['saldo_contas_pagar']:,.2f}"],
            ['Saldo Líquido', f"R$ {data['financeiro']['saldo_liquido']:,.2f}"],
            ['Vencidas (Receber)', f"R$ {data['financeiro']['contas_receber_vencidas']:,.2f}"],
            ['Vencidas (Pagar)', f"R$ {data['financeiro']['contas_pagar_vencidas']:,.2f}"]
        ]
        
        for item in fin_data:
            row += 1
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col, value=value)
                ExcelExporter.style_cell(cell)
        
        # CRM Section
        row += 2
        ws[f'A{row}'] = "CRM - PIPELINE DE VENDAS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FF6B35")
        
        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            ExcelExporter.style_header(cell)
        
        crm_data = [
            ['Oportunidades Abertas', data['crm']['oportunidades_abertas']],
            ['Valor do Pipeline', f"R$ {data['crm']['valor_pipeline']:,.2f}"]
        ]
        
        for item in crm_data:
            row += 1
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col, value=value)
                ExcelExporter.style_cell(cell)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
